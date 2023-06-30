from pickle import NONE
import interface

from openpyxl import *
from datetime import datetime

interf = interface.interface()

def main():
    nSensor = 4
    fileName = 'wristband/v7/adi_v3_1ADC_average.xlsx'
    gestures = ['down', 'up', 'thumb', 'little finger', 'stretch', 'fist', 'rest']
    workbook = load_workbook(fileName)
    start = datetime.now()
    mode = int(input("0: calibration, 1: random: "))
    if mode==0:
        worksheet = workbook.create_sheet("calibration")
    else:
        worksheet = workbook.create_sheet("random")
    title = ['gesture', 'start', 'end', 0, 1, 2, 3, 4, 5, 6, 7]
    for i in range(len(title)):
        worksheet.cell(row=1, column=i+1, value=title[i])
    row = 2
    while True:
        if mode==0:
            gesture = 4
        else:
            gesture = (row-2)%7
        _input = input("{}\tinput: ".format(gestures[gesture]))
        if _input == 'b':
            row -= 1
        elif _input == 'e':
            interf.end_process()
            workbook.save(fileName)
            workbook.close()
            break
        else:
            interf.write(str(1))
            worksheet.cell(row=row, column=1, value=gesture)
            worksheet.cell(row=row, column=2, value=str(datetime.now()-start))

            avg = [0, 0, 0, 0, 0, 0, 0, 0]
            for j in range(20):
                for k in range(nSensor):
                    btIn = float(interf.read())
                    while btIn == 0:
                        print(0)
                        btIn = float(interf.read())
                    avg[k] += btIn
            print(row-1, end='\t')
            for k in range(nSensor):
                avg[k] /= 20
                res = 300 * avg[k] / (5000 - avg[k] * 3)
                worksheet.cell(row=row, column=k+4, value=round(res, 2))
                workbook.save(fileName)
                # if not (k == 1 or k == 2 or k == 4 or k == 6):
                print(round(res, 2), end='\t')
            print()
            worksheet.cell(row=row, column=3, value=str(datetime.now()-start))
            row += 1

if __name__ == '__main__':
    main()